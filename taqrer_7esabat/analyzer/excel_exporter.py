"""
Professional Excel Export Generator for Django Dashboard Application.
Generates a 4-sheet .xlsx file with native Excel charts, RTL layout, 
corporate styling, conditional formatting, and transparent data quality flags.
"""

import io
from datetime import datetime
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.chart import (
    BarChart,
    LineChart,
    PieChart,
    DoughnutChart,
    ScatterChart,
    Reference,
    Series
)


class ExcelDashboardExporter:
    """Generates a professional 4-sheet Excel report from DataAnalyzer result and DataFrame."""

    def __init__(self, analyzer, result_data, filters_dict, original_filename):
        self.analyzer = analyzer
        self.df = analyzer.df if analyzer.df is not None else analyzer.original_df
        self.original_df = analyzer.original_df if analyzer.original_df is not None else self.df
        self.result = result_data
        self.filters = filters_dict or {}
        self.original_filename = original_filename or "dataset.xlsx"
        self.wb = openpyxl.Workbook()
        
        # Color Palette - Professional Corporate Slate / Indigo / Emerald Theme
        self.HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")  # Slate 800
        self.ACCENT_FILL = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")  # Indigo 600
        self.SUBHEADER_FILL = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")  # Slate 100
        self.ZEBRA_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")  # Slate 50
        
        # Severity Fills for Quality
        self.DANGER_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")  # Red 100
        self.WARNING_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")  # Amber 100
        self.SUCCESS_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")  # Emerald 100
        self.INFO_FILL = PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid")  # Sky 100
        
        # Fonts
        self.HEADER_FONT = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        self.TITLE_FONT = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
        self.SECTION_FONT = Font(name="Segoe UI", size=12, bold=True, color="1E293B")
        self.BOLD_FONT = Font(name="Segoe UI", size=10, bold=True, color="0F172A")
        self.DATA_FONT = Font(name="Segoe UI", size=10, color="334155")
        
        # Borders
        thin_border = Side(style='thin', color='CBD5E1')
        self.BORDER_ALL = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
        
        # Alignments
        self.ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.ALIGN_RIGHT = Alignment(horizontal='right', vertical='center', wrap_text=True)
        self.ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

    def generate(self):
        """Build the complete 4-sheet workbook and return bytes."""
        # Remove default sheet
        default_sheet = self.wb.active

        # Sheet 1: Summary / الملخص
        self._build_summary_sheet()

        # Sheet 2: Data Quality / جودة البيانات
        self._build_quality_sheet()

        # Sheet 3: Charts Data / بيانات الرسومات
        self._build_charts_sheet()

        # Sheet 4: Raw Data / البيانات الأصلية
        self._build_raw_data_sheet()

        if default_sheet and default_sheet in self.wb.worksheets:
            self.wb.remove(default_sheet)

        output = io.BytesIO()
        self.wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _setup_sheet(self, ws, title):
        """Set sheet title, RTL direction, and gridlines."""
        ws.title = title
        ws.views.sheetView[0].rightToLeft = True
        ws.sheet_properties.tabColor = "4F46E5"
        ws.views.sheetView[0].showGridLines = True

    # =========================================================
    # SHEET 1: SUMMARY / الملخص
    # =========================================================

    def _build_summary_sheet(self):
        ws = self.wb.create_sheet(title="الملخص")
        self._setup_sheet(ws, "الملخص")

        # Row 1: Header Banner
        ws.merge_cells("A1:E1")
        title_cell = ws["A1"]
        title_cell.value = "تقرير لوحة التحكم والتحليلات — تقرير حسابات"
        title_cell.font = self.TITLE_FONT
        title_cell.fill = self.HEADER_FILL
        title_cell.alignment = self.ALIGN_CENTER
        ws.row_dimensions[1].height = 40

        # Row 2-5: Metadata Block
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
        mode_str = "تحليل مخصص" if self.result.get('analysis_mode') == 'custom' else "تحليل تلقائي شامل"
        query_str = f" (الطلب: {self.result.get('custom_query')})" if self.result.get('custom_query') else ""

        # Filter info text
        active_filters = []
        if self.filters.get('date_from'):
            active_filters.append(f"من تاريخ: {self.filters['date_from']}")
        if self.filters.get('date_to'):
            active_filters.append(f"إلى تاريخ: {self.filters['date_to']}")
        if self.filters.get('category'):
            col = self.filters.get('category_column', '')
            active_filters.append(f"التصنيف: {self.filters['category']} ({col})")
        
        filter_str = " | ".join(active_filters) if active_filters else "جميع البيانات (بدون تصفية)"

        meta_rows = [
            ("اسم الملف الأصلي:", self.original_filename, "تاريخ التصدير:", now_str),
            ("نوع التحليل:", f"{mode_str}{query_str}", "إجمالي السجلات:", f"{len(self.df):,}"),
            ("الفلاتر النشطة:", filter_str, "", ""),
        ]

        for r_idx, row_data in enumerate(meta_rows, start=2):
            ws.cell(row=r_idx, column=1, value=row_data[0]).font = self.BOLD_FONT
            ws.cell(row=r_idx, column=1).alignment = self.ALIGN_RIGHT
            
            c2 = ws.cell(row=r_idx, column=2, value=row_data[1])
            c2.font = self.DATA_FONT
            c2.alignment = self.ALIGN_RIGHT
            
            if row_data[2]:
                ws.cell(row=r_idx, column=4, value=row_data[2]).font = self.BOLD_FONT
                ws.cell(row=r_idx, column=4).alignment = self.ALIGN_RIGHT
                c5 = ws.cell(row=r_idx, column=5, value=row_data[3])
                c5.font = self.DATA_FONT
                c5.alignment = self.ALIGN_RIGHT
            ws.row_dimensions[r_idx].height = 22

        # Row 6: Section Divider
        curr_row = 6
        ws.cell(row=curr_row, column=1, value="").fill = self.SUBHEADER_FILL
        curr_row += 1

        # Executive Summary Section
        ws.cell(row=curr_row, column=1, value="📌 الملخص التنفيذي للمحلل").font = self.SECTION_FONT
        ws.row_dimensions[curr_row].height = 25
        curr_row += 1

        narrative_text = self.result.get('narrative', 'لا يوجد ملخص متاح.')
        ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row+2, end_column=5)
        n_cell = ws.cell(row=curr_row, column=1, value=narrative_text)
        n_cell.font = Font(name="Segoe UI", size=11, color="1E293B", italic=True)
        n_cell.fill = self.ZEBRA_FILL
        n_cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        
        for r in range(curr_row, curr_row+3):
            for c in range(1, 6):
                cell = ws.cell(row=r, column=c)
                cell.border = self.BORDER_ALL
                if not cell.fill.start_color.rgb:
                    cell.fill = self.ZEBRA_FILL
        curr_row += 4

        # KPIs Section
        ws.cell(row=curr_row, column=1, value="📊 المؤشرات الرئيسية (KPIs)").font = self.SECTION_FONT
        ws.row_dimensions[curr_row].height = 25
        curr_row += 1

        headers = ["اسم المؤشر", "القيمة المعروضة", "القيمة الرقمية", "الاتجاه / الوصف", "نوع المؤشر"]
        ws.row_dimensions[curr_row].height = 26
        for c_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=curr_row, column=c_idx, value=h)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_ALL
        
        kpi_start_row = curr_row
        curr_row += 1

        kpis = self.result.get('kpis', [])
        for k_idx, kpi in enumerate(kpis):
            ws.row_dimensions[curr_row].height = 22
            row_fill = self.ZEBRA_FILL if k_idx % 2 == 1 else PatternFill(fill_type=None)

            c1 = ws.cell(row=curr_row, column=1, value=kpi.get('label', ''))
            c1.font = self.BOLD_FONT

            c2 = ws.cell(row=curr_row, column=2, value=kpi.get('value', ''))
            c2.font = self.BOLD_FONT

            raw_val = kpi.get('raw_value')
            c3 = ws.cell(row=curr_row, column=3, value=raw_val if raw_val is not None else "")
            if isinstance(raw_val, (int, float)):
                c3.number_format = "#,##0.00" if isinstance(raw_val, float) else "#,##0"

            trend_val = kpi.get('trend_value', 0)
            trend_dir = kpi.get('trend', '')
            subtitle = kpi.get('subtitle', '')
            trend_str = f"{'▲' if trend_dir == 'up' else '▼' if trend_dir == 'down' else ''} {trend_val}%" if trend_val else subtitle

            c4 = ws.cell(row=curr_row, column=4, value=trend_str)
            c5 = ws.cell(row=curr_row, column=5, value=kpi.get('type', ''))

            for c_idx in range(1, 6):
                cell = ws.cell(row=curr_row, column=c_idx)
                cell.font = self.DATA_FONT if c_idx not in (1, 2) else self.BOLD_FONT
                cell.border = self.BORDER_ALL
                cell.alignment = self.ALIGN_RIGHT if c_idx == 1 else self.ALIGN_CENTER
                if row_fill.fill_type:
                    cell.fill = row_fill

            curr_row += 1

        ws.freeze_panes = f"A{kpi_start_row + 1}"
        self._auto_fit_columns(ws)

    # =========================================================
    # SHEET 2: DATA QUALITY / جودة البيانات
    # =========================================================

    def _build_quality_sheet(self):
        ws = self.wb.create_sheet(title="جودة البيانات")
        self._setup_sheet(ws, "جودة البيانات")

        quality = self.result.get('quality', {})
        completeness = quality.get('completeness_score', 100.0)

        # Title
        ws.merge_cells("A1:E1")
        t = ws["A1"]
        t.value = "تقرير جودة البيانات وتدقيق الصحة"
        t.font = self.TITLE_FONT
        t.fill = self.HEADER_FILL
        t.alignment = self.ALIGN_CENTER
        ws.row_dimensions[1].height = 40

        # Quality Overview KPI boxes
        ws.cell(row=3, column=1, value="درجة اكتمال البيانات:").font = self.BOLD_FONT
        c_comp = ws.cell(row=3, column=2, value=completeness / 100.0)
        c_comp.font = Font(name="Segoe UI", size=12, bold=True, color="10B981" if completeness >= 90 else "EF4444")
        c_comp.number_format = "0.0%"

        ws.cell(row=3, column=4, value="إجمالي الصفوف المكررة:").font = self.BOLD_FONT
        ws.cell(row=3, column=5, value=quality.get('details', {}).get('duplicate_count', 0)).font = self.BOLD_FONT

        curr_row = 5
        ws.cell(row=curr_row, column=1, value="📋 تفاصيل الأعمدة والقيم المفقودة").font = self.SECTION_FONT
        ws.row_dimensions[curr_row].height = 25
        curr_row += 1

        # Table 1: Missing values per column
        headers1 = ["اسم العمود", "نوع العمود (المستنتج)", "عدد القيم المفقودة", "نسبة المفقودات (%)", "حالة العمود"]
        ws.row_dimensions[curr_row].height = 26
        for c_idx, h in enumerate(headers1, start=1):
            cell = ws.cell(row=curr_row, column=c_idx, value=h)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_ALL

        col_start_row = curr_row + 1
        curr_row += 1

        col_types = self.result.get('column_types', {})
        missing_details = quality.get('details', {}).get('missing_by_column', {})

        for col_name in self.df.columns:
            ws.row_dimensions[curr_row].height = 22
            col_type = col_types.get(col_name, 'text')
            
            m_info = missing_details.get(col_name, {})
            m_count = m_info.get('count', int(self.df[col_name].isnull().sum()))
            m_pct = m_info.get('percentage', round(m_count / len(self.df) * 100, 1) if len(self.df) > 0 else 0)

            status_str = "سليم" if m_count == 0 else ("تنبيه: مفقودات مرتفعة" if m_pct >= 15 else "مفقودات جزئية")

            ws.cell(row=curr_row, column=1, value=col_name).font = self.BOLD_FONT
            ws.cell(row=curr_row, column=2, value=col_type)
            
            c_cnt = ws.cell(row=curr_row, column=3, value=m_count)
            c_cnt.number_format = "#,##0"

            c_pct = ws.cell(row=curr_row, column=4, value=m_pct / 100.0)
            c_pct.number_format = "0.0%"

            ws.cell(row=curr_row, column=5, value=status_str)

            for c_idx in range(1, 6):
                cell = ws.cell(row=curr_row, column=c_idx)
                cell.font = self.DATA_FONT if c_idx != 1 else self.BOLD_FONT
                cell.border = self.BORDER_ALL
                cell.alignment = self.ALIGN_CENTER if c_idx != 1 else self.ALIGN_RIGHT

            curr_row += 1

        col_end_row = curr_row - 1

        # Conditional Formatting: Highlight column missing % > 15% with warning fill
        if col_end_row >= col_start_row:
            rule = CellIsRule(operator='greaterThan', formula=['0.15'], fill=self.DANGER_FILL, font=Font(color="991B1B", bold=True))
            ws.conditional_formatting.add(f"D{col_start_row}:D{col_end_row}", rule)

        curr_row += 2

        # Table 2: Data Quality Issues & Flags
        ws.cell(row=curr_row, column=1, value="⚠️ ملاحظات وتنبيهات جودة البيانات").font = self.SECTION_FONT
        ws.row_dimensions[curr_row].height = 25
        curr_row += 1

        headers2 = ["نوع الملاحظة", "درجة الأهمية", "تفاصيل الملاحظة", "العمود المتأثر"]
        ws.row_dimensions[curr_row].height = 26
        for c_idx, h in enumerate(headers2, start=1):
            cell = ws.cell(row=curr_row, column=c_idx, value=h)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_ALL
        curr_row += 1

        issues = quality.get('issues', [])
        for iss in issues:
            ws.row_dimensions[curr_row].height = 22
            sev = iss.get('severity', 'info')
            fill = self.DANGER_FILL if sev == 'danger' else (self.WARNING_FILL if sev == 'warning' else self.INFO_FILL)

            ws.cell(row=curr_row, column=1, value=iss.get('type', ''))
            
            c_sev = ws.cell(row=curr_row, column=2, value=sev.upper())
            c_sev.fill = fill
            c_sev.font = self.BOLD_FONT

            ws.cell(row=curr_row, column=3, value=iss.get('message', ''))
            ws.cell(row=curr_row, column=4, value=iss.get('affected_column') or "عام")

            for c_idx in range(1, 5):
                cell = ws.cell(row=curr_row, column=c_idx)
                cell.border = self.BORDER_ALL
                cell.alignment = self.ALIGN_RIGHT if c_idx == 3 else self.ALIGN_CENTER

            curr_row += 1

        ws.freeze_panes = "A6"
        self._auto_fit_columns(ws)

    # =========================================================
    # SHEET 3: CHARTS DATA & NATIVE EXCEL CHARTS
    # =========================================================

    def _build_charts_sheet(self):
        ws = self.wb.create_sheet(title="بيانات الرسومات")
        self._setup_sheet(ws, "بيانات الرسومات")

        # Title
        ws.merge_cells("A1:G1")
        t = ws["A1"]
        t.value = "البيانات التجميعية والرسومات البيانية التفاعلية"
        t.font = self.TITLE_FONT
        t.fill = self.HEADER_FILL
        t.alignment = self.ALIGN_CENTER
        ws.row_dimensions[1].height = 40

        charts = self.result.get('charts', [])
        curr_row = 3

        for ch_idx, chart_data in enumerate(charts, start=1):
            chart_type = chart_data.get('type', 'bar')
            chart_title = chart_data.get('title', f'رسم بياني {ch_idx}')
            data_spec = chart_data.get('data', {})
            labels = data_spec.get('labels', [])
            datasets = data_spec.get('datasets', [])

            if not datasets:
                continue

            # Chart Title Header
            ws.cell(row=curr_row, column=1, value=f"{ch_idx}. {chart_title}").font = self.SECTION_FONT
            ws.row_dimensions[curr_row].height = 25
            curr_row += 1

            table_start_row = curr_row

            # Build Data Table for this chart
            if chart_type == 'scatter':
                # Scatter data format
                ds = datasets[0]
                points = ds.get('data', [])
                ws.cell(row=curr_row, column=1, value="X").font = self.HEADER_FONT
                ws.cell(row=curr_row, column=1).fill = self.HEADER_FILL
                ws.cell(row=curr_row, column=2, value="Y").font = self.HEADER_FONT
                ws.cell(row=curr_row, column=2).fill = self.HEADER_FILL
                curr_row += 1

                for pt in points:
                    ws.cell(row=curr_row, column=1, value=pt.get('x', 0)).font = self.DATA_FONT
                    ws.cell(row=curr_row, column=2, value=pt.get('y', 0)).font = self.DATA_FONT
                    curr_row += 1
                table_end_row = curr_row - 1

                # Native Scatter Chart
                xl_chart = ScatterChart()
                xl_chart.title = chart_title
                xl_chart.style = 13
                xvalues = Reference(ws, min_col=1, min_row=table_start_row+1, max_row=table_end_row)
                yvalues = Reference(ws, min_col=2, min_row=table_start_row+1, max_row=table_end_row)
                series = Series(yvalues, xvalues, title_from_data=False)
                xl_chart.series.append(series)

            else:
                # Standard Bar / Doughnut / Line / Grouped Bar
                ws.cell(row=curr_row, column=1, value="الفئة / التاريخ").font = self.HEADER_FONT
                ws.cell(row=curr_row, column=1).fill = self.HEADER_FILL
                ws.cell(row=curr_row, column=1).alignment = self.ALIGN_CENTER

                for d_idx, ds in enumerate(datasets, start=2):
                    cell = ws.cell(row=curr_row, column=d_idx, value=ds.get('label', f'سلسلة {d_idx-1}'))
                    cell.font = self.HEADER_FONT
                    cell.fill = self.HEADER_FILL
                    cell.alignment = self.ALIGN_CENTER
                curr_row += 1

                for l_idx, lbl in enumerate(labels):
                    ws.cell(row=curr_row, column=1, value=str(lbl)).alignment = self.ALIGN_RIGHT
                    for d_idx, ds in enumerate(datasets, start=2):
                        vals = ds.get('data', [])
                        val = vals[l_idx] if l_idx < len(vals) else 0
                        cell = ws.cell(row=curr_row, column=d_idx, value=val if val is not None else 0)
                        if isinstance(val, (int, float)):
                            cell.number_format = "#,##0.00" if isinstance(val, float) else "#,##0"
                        cell.alignment = self.ALIGN_CENTER
                    curr_row += 1
                table_end_row = curr_row - 1

                # Apply borders to table
                for r in range(table_start_row, table_end_row + 1):
                    for c in range(1, len(datasets) + 2):
                        ws.cell(row=r, column=c).border = self.BORDER_ALL

                # Create Native Excel Chart
                num_datasets = len(datasets)
                cats = Reference(ws, min_col=1, min_row=table_start_row+1, max_row=table_end_row)
                data_ref = Reference(ws, min_col=2, max_col=1+num_datasets, min_row=table_start_row, max_row=table_end_row)

                if chart_type in ('doughnut', 'pie'):
                    xl_chart = DoughnutChart() if chart_type == 'doughnut' else PieChart()
                    xl_chart.title = chart_title
                    xl_chart.add_data(Reference(ws, min_col=2, min_row=table_start_row, max_row=table_end_row), titles_from_data=True)
                    xl_chart.set_categories(cats)
                elif chart_type == 'line':
                    xl_chart = LineChart()
                    xl_chart.title = chart_title
                    xl_chart.style = 12
                    xl_chart.add_data(data_ref, titles_from_data=True)
                    xl_chart.set_categories(cats)
                else:  # Bar chart
                    xl_chart = BarChart()
                    xl_chart.type = "col"
                    xl_chart.style = 10
                    xl_chart.title = chart_title
                    xl_chart.add_data(data_ref, titles_from_data=True)
                    xl_chart.set_categories(cats)
                    if chart_data.get('options', {}).get('stacked'):
                        xl_chart.grouping = "stacked"
                        xl_chart.overlap = 100

            xl_chart.width = 16
            xl_chart.height = 9.5
            # Place chart in column E next to its table
            chart_cell_pos = f"E{table_start_row}"
            ws.add_chart(xl_chart, chart_cell_pos)

            curr_row = max(curr_row + 2, table_start_row + 18)

        ws.freeze_panes = "A2"
        self._auto_fit_columns(ws)

    # =========================================================
    # SHEET 4: RAW DATA / البيانات الأصلية
    # =========================================================

    def _build_raw_data_sheet(self):
        ws = self.wb.create_sheet(title="البيانات الأصلية")
        self._setup_sheet(ws, "البيانات الأصلية")

        # Detect duplicated rows in original dataframe
        is_dup_series = self.df.duplicated(keep=False)
        is_null_series = self.df.isnull().any(axis=1)

        headers = ["حالة البيانات"] + list(self.df.columns)

        # Header Row
        ws.row_dimensions[1].height = 28
        for c_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c_idx, value=str(h))
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_ALL

        # Data Rows
        for r_idx, (df_idx, row) in enumerate(self.df.iterrows(), start=2):
            ws.row_dimensions[r_idx].height = 20
            
            # Row Status Flag (never silently remove!)
            is_dup = is_dup_series.loc[df_idx] if df_idx in is_dup_series.index else False
            is_null = is_null_series.loc[df_idx] if df_idx in is_null_series.index else False

            if is_dup:
                status_str = "صف مكرر"
                status_fill = self.WARNING_FILL
            elif is_null:
                status_str = "بيانات مفقودة"
                status_fill = self.INFO_FILL
            else:
                status_str = "سجل نظيف"
                status_fill = self.SUCCESS_FILL

            c_status = ws.cell(row=r_idx, column=1, value=status_str)
            c_status.font = self.BOLD_FONT
            c_status.fill = status_fill
            c_status.alignment = self.ALIGN_CENTER
            c_status.border = self.BORDER_ALL

            # Write DataFrame columns
            for c_idx, col_name in enumerate(self.df.columns, start=2):
                val = row[col_name]
                cell = ws.cell(row=r_idx, column=c_idx)

                if pd.isna(val):
                    cell.value = ""
                elif isinstance(val, (int, float, np.integer, np.floating)):
                    cell.value = float(val) if isinstance(val, (float, np.floating)) else int(val)
                    cell.number_format = "#,##0.00" if isinstance(val, (float, np.floating)) else "#,##0"
                elif isinstance(val, (datetime, pd.Timestamp)):
                    cell.value = val.strftime("%Y-%m-%d")
                else:
                    cell.value = str(val)

                cell.font = self.DATA_FONT
                cell.border = self.BORDER_ALL
                cell.alignment = self.ALIGN_LEFT if c_idx == 2 else self.ALIGN_CENTER

        ws.freeze_panes = "A2"
        self._auto_fit_columns(ws)

    # =========================================================
    # HELPERS
    # =========================================================

    def _auto_fit_columns(self, ws):
        """Auto-adjust column widths with sensible padding and max limit."""
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                # Don't size based on merged title row 1
                if cell.row == 1 and cell.coordinate != "A1":
                    continue
                if cell.value:
                    val_str = str(cell.value)
                    if len(val_str) < 100:  # Skip long paragraphs for width calc
                        max_len = max(max_len, len(val_str))

            # Apply width with padding
            adjusted_width = max(max_len + 4, 12)
            ws.column_dimensions[col_letter].width = min(adjusted_width, 45)
